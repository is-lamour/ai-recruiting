from fastapi import FastAPI, HTTPException, BackgroundTasks, Request, Body
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from typing import Optional
import json
import io
import sqlite3
import threading
import requests as http_requests
from pathlib import Path
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import init_db, get_db, row_to_dict
from ai import screen_resume, summarize_vacancy, generate_boolean_search, generate_metrics

# in-memory progress store: vacancy_id → {total, done, errors, running}
_rescreen_progress: dict[int, dict] = {}
_rescreen_lock = threading.Lock()

app = FastAPI(title="HH Auto Screening")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

STATIC_DIR = Path(__file__).parent / "static"


@app.on_event("startup")
def startup():
    init_db()


# ── Static / Dashboard ────────────────────────────────────────────────────────

app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


@app.get("/")
def root():
    return FileResponse(str(STATIC_DIR / "index.html"))


@app.get("/health")
def health():
    return {"status": "ok"}


# ── Schemas ───────────────────────────────────────────────────────────────────

class VacancyCreate(BaseModel):
    title: str
    description: str
    requirements: Optional[str] = None  # если передано — используется вместо AI-генерации


class CandidateScreen(BaseModel):
    vacancy_id: int
    name: Optional[str] = None
    hh_url: str
    resume_text: str


class StatusUpdate(BaseModel):
    status: str  # new | to_reject | to_huntflow | rejected | huntflow_sent


class BulkAction(BaseModel):
    ids: list
    action: str  # "to_reject" | "delete"


class RescreenRequest(BaseModel):
    candidate_ids: Optional[list[int]] = None  # None = все кандидаты вакансии


# ── Background tasks ──────────────────────────────────────────────────────────

def _get_gemini_key(request: Request) -> Optional[str]:
    return request.headers.get("X-Gemini-Key") or None


def _generate_summary_bg(vacancy_id: int, description: str, api_key: Optional[str] = None):
    """Генерирует summary вакансии в фоне и сохраняет в БД только если requirements пусто."""
    try:
        requirements = summarize_vacancy(description, api_key=api_key)
        conn = get_db()
        try:
            conn.execute(
                "UPDATE vacancies SET requirements = ? WHERE id = ? AND (requirements IS NULL OR requirements = '')",
                (requirements, vacancy_id),
            )
            conn.commit()
        finally:
            conn.close()
    except Exception as e:
        print(f"[summary bg error] vacancy_id={vacancy_id}: {e}")


def _generate_metrics_bg(vacancy_id: int, description: str, api_key: Optional[str] = None):
    """Генерирует метрики вакансии в фоне и сохраняет JSON в БД."""
    try:
        metrics = generate_metrics(description, api_key=api_key)
        conn = get_db()
        try:
            conn.execute(
                "UPDATE vacancies SET metrics = ? WHERE id = ?",
                (json.dumps(metrics, ensure_ascii=False), vacancy_id),
            )
            conn.commit()
        finally:
            conn.close()
    except Exception as e:
        print(f"[metrics bg error] vacancy_id={vacancy_id}: {e}")


def _generate_boolean_search_bg(vacancy_id: int, description: str, api_key: Optional[str] = None):
    """Генерирует Boolean search в фоне и сохраняет JSON в БД."""
    try:
        result = generate_boolean_search(description, api_key=api_key)
        conn = get_db()
        try:
            conn.execute(
                "UPDATE vacancies SET boolean_search = ? WHERE id = ?",
                (json.dumps(result, ensure_ascii=False), vacancy_id),
            )
            conn.commit()
        finally:
            conn.close()
    except Exception as e:
        print(f"[boolean_search bg error] vacancy_id={vacancy_id}: {e}")


# ── Vacancies ─────────────────────────────────────────────────────────────────

@app.get("/api/vacancies")
def list_vacancies():
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT * FROM vacancies ORDER BY created_at DESC"
        ).fetchall()
        return [row_to_dict(r) for r in rows]
    finally:
        conn.close()


@app.post("/api/vacancies")
def create_vacancy(request: Request, data: VacancyCreate, background_tasks: BackgroundTasks):
    manual_req = data.requirements.strip() if data.requirements and data.requirements.strip() else None
    api_key = _get_gemini_key(request)
    conn = get_db()
    try:
        cur = conn.execute(
            "INSERT INTO vacancies (title, description, requirements) VALUES (?, ?, ?)",
            (data.title, data.description, manual_req or ""),
        )
        conn.commit()
        vacancy_id = cur.lastrowid
        if not manual_req:
            background_tasks.add_task(_generate_summary_bg, vacancy_id, data.description, api_key)
        background_tasks.add_task(_generate_metrics_bg, vacancy_id, data.description, api_key)
        row = conn.execute("SELECT * FROM vacancies WHERE id = ?", (vacancy_id,)).fetchone()
        return row_to_dict(row)
    finally:
        conn.close()


@app.put("/api/vacancies/{vacancy_id}")
def update_vacancy(vacancy_id: int, request: Request, data: VacancyCreate, background_tasks: BackgroundTasks):
    manual_req = data.requirements.strip() if data.requirements and data.requirements.strip() else None
    api_key = _get_gemini_key(request)
    conn = get_db()
    try:
        conn.execute(
            "UPDATE vacancies SET title = ?, description = ?, requirements = ? WHERE id = ?",
            (data.title, data.description, manual_req or "", vacancy_id),
        )
        conn.commit()
        if not manual_req:
            background_tasks.add_task(_generate_summary_bg, vacancy_id, data.description, api_key)
        background_tasks.add_task(_generate_metrics_bg, vacancy_id, data.description, api_key)
        row = conn.execute("SELECT * FROM vacancies WHERE id = ?", (vacancy_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Не найдено")
        return row_to_dict(row)
    finally:
        conn.close()


@app.delete("/api/vacancies/{vacancy_id}")
def delete_vacancy(vacancy_id: int):
    conn = get_db()
    try:
        conn.execute("DELETE FROM candidates WHERE vacancy_id = ?", (vacancy_id,))
        conn.execute("DELETE FROM vacancies WHERE id = ?", (vacancy_id,))
        conn.commit()
        return {"status": "deleted"}
    finally:
        conn.close()


@app.post("/api/vacancies/{vacancy_id}/generate-boolean")
def generate_boolean_for_vacancy(vacancy_id: int, request: Request, background_tasks: BackgroundTasks):
    """Запускает генерацию Boolean search для вакансии в фоне."""
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM vacancies WHERE id = ?", (vacancy_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Вакансия не найдена")
        vacancy = row_to_dict(row)
    finally:
        conn.close()

    # Сбрасываем текущий результат, чтобы фронт знал — генерация началась
    conn = get_db()
    try:
        conn.execute("UPDATE vacancies SET boolean_search = '' WHERE id = ?", (vacancy_id,))
        conn.commit()
    finally:
        conn.close()

    api_key = _get_gemini_key(request)
    background_tasks.add_task(
        _generate_boolean_search_bg,
        vacancy_id,
        vacancy["description"],
        api_key,
    )
    return {"status": "generating"}


@app.get("/api/vacancies/{vacancy_id}/metrics")
def get_metrics(vacancy_id: int):
    conn = get_db()
    try:
        row = conn.execute("SELECT metrics FROM vacancies WHERE id = ?", (vacancy_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Вакансия не найдена")
        raw = row["metrics"] or "[]"
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            return []
    finally:
        conn.close()


@app.put("/api/vacancies/{vacancy_id}/metrics")
def update_metrics(vacancy_id: int, metrics: list = Body(...)):
    # Валидируем и нормализуем
    clean = []
    for item in metrics:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name", "")).strip()
        if not name:
            continue
        try:
            weight = float(item.get("weight", 5.0))
            weight = max(0.0, min(10.0, round(weight * 2) / 2))
        except (ValueError, TypeError):
            weight = 5.0
        clean.append({"name": name, "weight": weight})
    conn = get_db()
    try:
        conn.execute(
            "UPDATE vacancies SET metrics = ? WHERE id = ?",
            (json.dumps(clean, ensure_ascii=False), vacancy_id),
        )
        conn.commit()
        return clean
    finally:
        conn.close()


@app.post("/api/vacancies/{vacancy_id}/generate-metrics")
def generate_metrics_for_vacancy(vacancy_id: int, request: Request, background_tasks: BackgroundTasks):
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM vacancies WHERE id = ?", (vacancy_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Вакансия не найдена")
        vacancy = row_to_dict(row)
    finally:
        conn.close()

    conn = get_db()
    try:
        conn.execute("UPDATE vacancies SET metrics = '[]' WHERE id = ?", (vacancy_id,))
        conn.commit()
    finally:
        conn.close()

    api_key = _get_gemini_key(request)
    background_tasks.add_task(_generate_metrics_bg, vacancy_id, vacancy["description"], api_key)
    return {"status": "generating"}


@app.get("/api/vacancies/{vacancy_id}/screened-urls")
def screened_urls(vacancy_id: int):
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT hh_url FROM candidates WHERE vacancy_id = ?", (vacancy_id,)
        ).fetchall()
        return [row["hh_url"] for row in rows]
    finally:
        conn.close()


# ── Parse HH vacancy URL ──────────────────────────────────────────────────────

@app.get("/api/parse-hh-vacancy")
def parse_hh_vacancy(url: str):
    """Загружает страницу вакансии HH и возвращает title + description."""
    url = url.strip().strip("<>")
    try:
        resp = http_requests.get(url, timeout=15, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120",
            "Accept-Language": "ru-RU,ru;q=0.9",
        })
        resp.raise_for_status()
        resp.encoding = "utf-8"
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Не удалось загрузить страницу: {e}")

    soup = BeautifulSoup(resp.text, "html.parser")

    # Заголовок
    title_el = (
        soup.find(attrs={"data-qa": "vacancy-title"}) or
        soup.find("h1")
    )
    title = title_el.get_text(" ", strip=True) if title_el else ""
    # Убираем подзаголовки вроде "(архив 7 мая 2025)"
    if title_el:
        for sub in title_el.find_all(["div", "span"]):
            sub.decompose()
        title = title_el.get_text(strip=True)

    # Описание — несколько вариантов селекторов
    desc_el = (
        soup.find(attrs={"data-qa": "vacancy-description"}) or
        soup.find("div", class_=lambda c: c and "vacancy-description" in " ".join(c)) or
        soup.find(attrs={"data-qa": "vacancy-branded-description"}) or
        soup.find("div", class_=lambda c: c and "brandingDescription" in " ".join(c if c else []))
    )
    description = desc_el.get_text(separator="\n", strip=True) if desc_el else ""

    print(f"[parse-hh] title={repr(title[:60])} desc_len={len(description)}")

    if not title:
        raise HTTPException(status_code=422, detail="Не удалось найти заголовок вакансии")
    if not description:
        raise HTTPException(status_code=422, detail="Не удалось найти описание вакансии. Попробуйте вставить текст вручную.")

    return {"title": title, "description": description}


# ── Candidate deserializer ───────────────────────────────────────────────────

def _parse_candidate(row) -> dict:
    d = row_to_dict(row)
    d["questions"]       = json.loads(d.get("questions")       or "[]")
    d["pros"]            = json.loads(d.get("pros")            or "[]")
    d["cons"]            = json.loads(d.get("cons")            or "[]")
    d["score_breakdown"] = json.loads(d.get("score_breakdown") or "[]")
    return d


# ── Screening ─────────────────────────────────────────────────────────────────

@app.post("/api/screen")
def screen_candidate(request: Request, data: CandidateScreen):
    conn = get_db()
    try:
        vacancy = conn.execute(
            "SELECT * FROM vacancies WHERE id = ?", (data.vacancy_id,)
        ).fetchone()
        if not vacancy:
            raise HTTPException(status_code=404, detail="Вакансия не найдена")

        requirements = vacancy["requirements"] or vacancy["description"][:500]
        raw_metrics = vacancy["metrics"] if "metrics" in vacancy.keys() else "[]"
        try:
            metrics = json.loads(raw_metrics or "[]") if isinstance(raw_metrics, str) else (raw_metrics or [])
        except Exception:
            metrics = []
        result = screen_resume(requirements, data.resume_text, api_key=_get_gemini_key(request), metrics=metrics or None)

        try:
            cur = conn.execute(
                """INSERT INTO candidates
                   (vacancy_id, name, hh_url, resume_text, score, category, ai_comment,
                    questions, summary, pros, cons, score_breakdown, is_trashed)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0)""",
                (
                    data.vacancy_id,
                    data.name,
                    data.hh_url,
                    data.resume_text,
                    result["score"],
                    result["category"],
                    result["comment"],
                    json.dumps(result["questions"],       ensure_ascii=False),
                    result.get("summary", ""),
                    json.dumps(result.get("pros", []),    ensure_ascii=False),
                    json.dumps(result.get("cons", []),    ensure_ascii=False),
                    json.dumps(result.get("score_breakdown", []), ensure_ascii=False),
                ),
            )
            conn.commit()
            row = conn.execute(
                "SELECT * FROM candidates WHERE id = ?", (cur.lastrowid,)
            ).fetchone()
        except sqlite3.IntegrityError:
            # Параллельный запрос уже вставил эту пару (vacancy_id, hh_url) — возвращаем существующую
            row = conn.execute(
                "SELECT * FROM candidates WHERE hh_url = ? AND vacancy_id = ?",
                (data.hh_url, data.vacancy_id),
            ).fetchone()

        return _parse_candidate(row)
    finally:
        conn.close()


# ── Candidates ────────────────────────────────────────────────────────────────

@app.get("/api/vacancies/{vacancy_id}/candidates")
def list_candidates(vacancy_id: int, category: Optional[str] = None, trashed: bool = False):
    conn = get_db()
    try:
        query = "SELECT * FROM candidates WHERE vacancy_id = ? AND COALESCE(is_trashed, 0) = ?"
        params: list = [vacancy_id, 1 if trashed else 0]
        if category:
            query += " AND category = ?"
            params.append(category)
        query += " ORDER BY score DESC"
        rows = conn.execute(query, params).fetchall()
        return [_parse_candidate(r) for r in rows]
    finally:
        conn.close()


@app.patch("/api/candidates/{candidate_id}")
def update_status(candidate_id: int, data: StatusUpdate):
    conn = get_db()
    try:
        conn.execute(
            "UPDATE candidates SET status = ? WHERE id = ?",
            (data.status, candidate_id),
        )
        conn.commit()
        row = conn.execute(
            "SELECT * FROM candidates WHERE id = ?", (candidate_id,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Кандидат не найден")
        return _parse_candidate(row)
    finally:
        conn.close()


@app.delete("/api/candidates/{candidate_id}")
def delete_candidate(candidate_id: int):
    conn = get_db()
    try:
        conn.execute("UPDATE candidates SET is_trashed = 1 WHERE id = ?", (candidate_id,))
        conn.commit()
        return {"status": "trashed"}
    finally:
        conn.close()


@app.post("/api/candidates/{candidate_id}/restore")
def restore_candidate(candidate_id: int):
    conn = get_db()
    try:
        conn.execute("UPDATE candidates SET is_trashed = 0 WHERE id = ?", (candidate_id,))
        conn.commit()
        return {"status": "restored"}
    finally:
        conn.close()


@app.delete("/api/candidates/{candidate_id}/permanent")
def delete_candidate_permanent(candidate_id: int):
    conn = get_db()
    try:
        conn.execute("DELETE FROM candidates WHERE id = ?", (candidate_id,))
        conn.commit()
        return {"status": "deleted"}
    finally:
        conn.close()


@app.post("/api/candidates/bulk")
def bulk_action(data: BulkAction):
    if not data.ids:
        return {"affected": 0}
    conn = get_db()
    try:
        placeholders = ",".join("?" * len(data.ids))
        if data.action == "delete":
            conn.execute(
                f"UPDATE candidates SET is_trashed = 1 WHERE id IN ({placeholders})",
                data.ids,
            )
        elif data.action == "to_reject":
            conn.execute(
                f"UPDATE candidates SET status = 'to_reject' WHERE id IN ({placeholders})",
                data.ids,
            )
        conn.commit()
        return {"affected": len(data.ids)}
    finally:
        conn.close()


@app.delete("/api/vacancies/{vacancy_id}/trash")
def empty_trash(vacancy_id: int):
    conn = get_db()
    try:
        conn.execute(
            "DELETE FROM candidates WHERE vacancy_id = ? AND is_trashed = 1",
            (vacancy_id,),
        )
        conn.commit()
        return {"status": "emptied"}
    finally:
        conn.close()


# ── Export ────────────────────────────────────────────────────────────────────

@app.get("/api/vacancies/{vacancy_id}/export")
def export_candidates(vacancy_id: int):
    conn = get_db()
    try:
        vacancy_row = conn.execute("SELECT * FROM vacancies WHERE id = ?", (vacancy_id,)).fetchone()
        if not vacancy_row:
            raise HTTPException(status_code=404, detail="Вакансия не найдена")
        vacancy = row_to_dict(vacancy_row)

        rows = conn.execute(
            "SELECT * FROM candidates WHERE vacancy_id = ? ORDER BY score DESC", (vacancy_id,)
        ).fetchall()
        candidates = [_parse_candidate(r) for r in rows]
    finally:
        conn.close()

    # ── Стили ─────────────────────────────────────────────────────────────────
    FILLS = {
        "suitable": PatternFill("solid", fgColor="DCFCE7"),
        "consider":  PatternFill("solid", fgColor="FEF3C7"),
        "reject":    PatternFill("solid", fgColor="FEE2E2"),
        "pending":   PatternFill("solid", fgColor="F3F4F6"),
    }
    SCORE_FONT = {
        "suitable": Font(bold=True, color="16A34A"),
        "consider":  Font(bold=True, color="D97706"),
        "reject":    Font(bold=True, color="DC2626"),
        "pending":   Font(bold=True, color="6B7280"),
    }
    HEADER_FILL = PatternFill("solid", fgColor="1E293B")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    SECTION_FILL = PatternFill("solid", fgColor="334155")
    WRAP = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin", color="E2E8F0")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    CAT_LABELS = {"suitable": "Подходит", "consider": "Подумать", "reject": "Отказ", "pending": "Ожидание"}
    STATUS_LABELS = {
        "new": "", "to_reject": "На отказ", "to_huntflow": "В Huntflow",
        "rejected": "Отказ отправлен", "huntflow_sent": "Отправлен в HF",
    }

    # 13 колонок; резюме — последняя
    COLUMNS = [
        ("Балл",                 8),
        ("Категория",           12),
        ("Имя",                 24),
        ("Статус",              14),
        ("Дата скрининга",      16),
        ("Краткое резюме",      40),
        ("Плюсы",               36),
        ("Минусы",              36),
        ("Комментарий AI",      44),
        ("Оценка по критериям", 50),
        ("Вопросы по пробелам", 50),
        ("Ссылка",              36),
        ("Резюме",              80),
    ]
    NCOLS = len(COLUMNS)
    last_col_letter = get_column_letter(NCOLS)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = vacancy["title"][:30]

    def merge_row(row_num, value, font, height, bg=None):
        ws.merge_cells(f"A{row_num}:{last_col_letter}{row_num}")
        cell = ws[f"A{row_num}"]
        cell.value = value
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if bg:
            cell.fill = bg
        ws.row_dimensions[row_num].height = height

    current_row = 1

    # ── Строка 1: название вакансии ───────────────────────────────────────────
    merge_row(
        current_row,
        vacancy["title"],
        Font(bold=True, size=13, color="FFFFFF"),
        22,
        bg=PatternFill("solid", fgColor="1E293B"),
    )
    current_row += 1

    # ── Строка 2: описание вакансии ───────────────────────────────────────────
    description = vacancy.get("description", "")
    if description:
        desc_lines = max(3, min(len(description) // 120 + 1, 15))
        merge_row(
            current_row,
            description,
            Font(size=10, color="1E293B"),
            desc_lines * 14,
        )
        current_row += 1

    # ── Строка 3: требования (AI-сжатие) ─────────────────────────────────────
    requirements = vacancy.get("requirements", "")
    if requirements and requirements != description:
        merge_row(
            current_row,
            f"Требования: {requirements}",
            Font(size=10, italic=True, color="475569"),
            max(20, min(len(requirements) // 120 * 14 + 14, 80)),
        )
        current_row += 1

    # ── Строка 4: метрики ─────────────────────────────────────────────────────
    try:
        metrics = json.loads(vacancy.get("metrics") or "[]")
    except Exception:
        metrics = []
    if metrics:
        metrics_parts = [f"{m['name']} (вес {m['weight']})" for m in metrics if m.get("name")]
        merge_row(
            current_row,
            "Метрики оценки:  " + "  |  ".join(metrics_parts),
            Font(size=10, bold=True, color="FFFFFF"),
            18,
            bg=SECTION_FILL,
        )
        current_row += 1

    header_row = current_row

    # ── Заголовки столбцов ────────────────────────────────────────────────────
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[header_row].height = 22

    # ── Строки кандидатов ────────────────────────────────────────────────────
    for c in candidates:
        cat = c.get("category") or "pending"
        fill = FILLS.get(cat, FILLS["pending"])
        score_font = SCORE_FONT.get(cat, SCORE_FONT["pending"])

        questions_text  = "\n".join(f"• {q}" for q in c["questions"]) if c["questions"] else ""
        pros_text       = "\n".join(f"+ {p}" for p in c.get("pros", [])) if c.get("pros") else ""
        cons_text       = "\n".join(f"- {m}" for m in c.get("cons", [])) if c.get("cons") else ""

        breakdown = c.get("score_breakdown") or []
        breakdown_text = "\n".join(
            f"{b['criterion']}: {b['score']}/10  (вес {b['weight']})  — {b.get('note', '')}"
            for b in breakdown
        ) if breakdown else ""

        screened_at = str(c.get("created_at") or "")[:16].replace("T", " ")

        data_row = [
            c.get("score"),                                    # 1
            CAT_LABELS.get(cat, cat),                         # 2
            c.get("name") or "Кандидат",                      # 3
            STATUS_LABELS.get(c.get("status", "new"), ""),    # 4
            screened_at,                                       # 5
            c.get("summary") or "",                           # 6
            pros_text,                                         # 7
            cons_text,                                         # 8
            c.get("ai_comment") or "",                        # 9
            breakdown_text,                                    # 10
            questions_text,                                    # 11
            c.get("hh_url") or "",                            # 12
            c.get("resume_text") or "",                       # 13  — резюме в конце
        ]

        r = ws.max_row + 1
        for col_idx, value in enumerate(data_row, start=1):
            cell = ws.cell(row=r, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = WRAP
            cell.border = BORDER
            if col_idx == 1:
                cell.font = score_font
                cell.alignment = Alignment(horizontal="center", vertical="top")
            if col_idx == 12 and value:  # ссылка
                cell.hyperlink = value
                cell.font = Font(color="2563EB", underline="single")
            if col_idx == 13:  # резюме — серый текст, чуть меньше
                cell.font = Font(size=9, color="64748B")

        # Высота строки по самой «высокой» колонке (резюме не учитываем — слишком длинный)
        max_lines = max(
            len(str(data_row[5] or "").split("\n")),   # краткое резюме
            len(str(data_row[6] or "").split("\n")),   # плюсы
            len(str(data_row[7] or "").split("\n")),   # минусы
            len(str(data_row[8] or "").split("\n")),   # комментарий
            len(str(data_row[9] or "").split("\n")),   # breakdown
            len(str(data_row[10] or "").split("\n")),  # вопросы
            1,
        )
        ws.row_dimensions[r].height = max(18, min(max_lines * 15, 150))

    # ── Закрепить шапку ───────────────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # ── Отдать файл ───────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"candidates_{vacancy_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Stats ─────────────────────────────────────────────────────────────────────

@app.get("/api/vacancies/{vacancy_id}/stats")
def get_stats(vacancy_id: int):
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT category, COUNT(*) as cnt FROM candidates WHERE vacancy_id = ? AND COALESCE(is_trashed, 0) = 0 GROUP BY category",
            (vacancy_id,),
        ).fetchall()
        by_cat = {"suitable": 0, "consider": 0, "reject": 0, "pending": 0}
        total = 0
        for row in rows:
            by_cat[row["category"]] = row["cnt"]
            total += row["cnt"]
        trashed = conn.execute(
            "SELECT COUNT(*) as cnt FROM candidates WHERE vacancy_id = ? AND COALESCE(is_trashed, 0) = 1",
            (vacancy_id,),
        ).fetchone()["cnt"]
        return {"total": total, **by_cat, "trashed": trashed}
    finally:
        conn.close()


# ── Pending actions (для расширения) ─────────────────────────────────────────

@app.get("/api/vacancies/{vacancy_id}/pending-actions")
def pending_actions(vacancy_id: int):
    conn = get_db()
    try:
        to_reject = conn.execute(
            "SELECT id, hh_url, name FROM candidates WHERE vacancy_id = ? AND status = 'to_reject'",
            (vacancy_id,),
        ).fetchall()
        to_huntflow = conn.execute(
            "SELECT id, hh_url, name FROM candidates WHERE vacancy_id = ? AND status = 'to_huntflow'",
            (vacancy_id,),
        ).fetchall()
        return {
            "to_reject": [row_to_dict(r) for r in to_reject],
            "to_huntflow": [row_to_dict(r) for r in to_huntflow],
        }
    finally:
        conn.close()


@app.post("/api/vacancies/{vacancy_id}/mark-for-reject")
def mark_for_reject(vacancy_id: int):
    conn = get_db()
    try:
        conn.execute(
            "UPDATE candidates SET status = 'to_reject' WHERE vacancy_id = ? AND category = 'reject' AND status = 'new'",
            (vacancy_id,),
        )
        conn.commit()
        count = conn.execute(
            "SELECT COUNT(*) FROM candidates WHERE vacancy_id = ? AND status = 'to_reject'",
            (vacancy_id,),
        ).fetchone()[0]
        return {"queued": count}
    finally:
        conn.close()


@app.get("/api/vacancies/{vacancy_id}/rescreen/status")
def rescreen_status(vacancy_id: int):
    with _rescreen_lock:
        prog = _rescreen_progress.get(vacancy_id)
    if not prog:
        return {"running": False, "total": 0, "done": 0, "errors": 0}
    return prog


def _rescreen_bg(vacancy_id: int, candidate_ids: list[int], api_key: Optional[str]):
    conn = get_db()
    try:
        vacancy_row = conn.execute("SELECT * FROM vacancies WHERE id = ?", (vacancy_id,)).fetchone()
        if not vacancy_row:
            return
        vacancy = row_to_dict(vacancy_row)
        requirements = vacancy["requirements"] or vacancy["description"][:500]
        try:
            metrics = json.loads(vacancy.get("metrics") or "[]")
        except Exception:
            metrics = []

        rows = conn.execute(
            f"SELECT id, resume_text FROM candidates WHERE id IN ({','.join('?' * len(candidate_ids))}) AND vacancy_id = ?",
            (*candidate_ids, vacancy_id),
        ).fetchall()
    finally:
        conn.close()

    total = len(rows)
    with _rescreen_lock:
        _rescreen_progress[vacancy_id] = {"running": True, "total": total, "done": 0, "errors": 0}

    for row in rows:
        cid = row["id"]
        resume_text = row["resume_text"] or ""
        try:
            result = screen_resume(requirements, resume_text, api_key=api_key, metrics=metrics or None)
            conn = get_db()
            try:
                conn.execute(
                    """UPDATE candidates
                       SET score=?, category=?, ai_comment=?, questions=?, summary=?,
                           pros=?, cons=?, score_breakdown=?
                       WHERE id=?""",
                    (
                        result["score"],
                        result["category"],
                        result["comment"],
                        json.dumps(result["questions"],            ensure_ascii=False),
                        result.get("summary", ""),
                        json.dumps(result.get("pros", []),         ensure_ascii=False),
                        json.dumps(result.get("cons", []),         ensure_ascii=False),
                        json.dumps(result.get("score_breakdown", []), ensure_ascii=False),
                        cid,
                    ),
                )
                conn.commit()
            finally:
                conn.close()
            with _rescreen_lock:
                _rescreen_progress[vacancy_id]["done"] += 1
        except Exception as e:
            print(f"[rescreen] candidate {cid} error: {e}")
            with _rescreen_lock:
                _rescreen_progress[vacancy_id]["errors"] += 1
                _rescreen_progress[vacancy_id]["done"] += 1

    with _rescreen_lock:
        _rescreen_progress[vacancy_id]["running"] = False


@app.post("/api/vacancies/{vacancy_id}/rescreen")
def rescreen_candidates(vacancy_id: int, request: Request, data: RescreenRequest, background_tasks: BackgroundTasks):
    conn = get_db()
    try:
        if not conn.execute("SELECT 1 FROM vacancies WHERE id = ?", (vacancy_id,)).fetchone():
            raise HTTPException(status_code=404, detail="Вакансия не найдена")

        if data.candidate_ids:
            candidate_ids = data.candidate_ids
        else:
            rows = conn.execute(
                "SELECT id FROM candidates WHERE vacancy_id = ? AND COALESCE(is_trashed, 0) = 0",
                (vacancy_id,),
            ).fetchall()
            candidate_ids = [r["id"] for r in rows]
    finally:
        conn.close()

    if not candidate_ids:
        return {"status": "nothing_to_rescreen", "total": 0}

    with _rescreen_lock:
        if _rescreen_progress.get(vacancy_id, {}).get("running"):
            raise HTTPException(status_code=409, detail="Перескрининг уже запущен")

    api_key = _get_gemini_key(request)
    background_tasks.add_task(_rescreen_bg, vacancy_id, candidate_ids, api_key)
    return {"status": "started", "total": len(candidate_ids)}


@app.post("/api/vacancies/{vacancy_id}/mark-for-huntflow")
def mark_for_huntflow(vacancy_id: int):
    conn = get_db()
    try:
        conn.execute(
            "UPDATE candidates SET status = 'to_huntflow' WHERE vacancy_id = ? AND category = 'suitable' AND status = 'new'",
            (vacancy_id,),
        )
        conn.commit()
        count = conn.execute(
            "SELECT COUNT(*) FROM candidates WHERE vacancy_id = ? AND status = 'to_huntflow'",
            (vacancy_id,),
        ).fetchone()[0]
        return {"queued": count}
    finally:
        conn.close()
